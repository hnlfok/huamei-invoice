#!/usr/bin/env python3
"""华美物流发货单生成器 - Android/iOS App"""
import re, io, os, sys
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Kivy UI (支持Android/iOS)
try:
    import kivy
    from kivy.app import App
    from kivy.uix.boxlayout import BoxLayout
    from kivy.uix.gridlayout import GridLayout
    from kivy.uix.label import Label
    from kivy.uix.textinput import TextInput
    from kivy.uix.button import Button
    from kivy.uix.filechooser import FileChooserIconView
    from kivy.uix.scrollview import ScrollView
    from kivy.core.window import Window
    from kivy.core.text import LabelBase
    from kivy.utils import platform
    HAVE_KIVY = True
except ImportError:
    HAVE_KIVY = False

TEMPLATE_SHEET = "重庆华美物流有限公司重庆有研专用发货单"
ITEMS_START_ROW, ITEMS_END_ROW = 7, 14

def build_sheet(ws, recipient_name, recipient_phone, company_name,
                address, items, waybill_num, ship_date):
    ws.row_dimensions[1].height = 30
    ws["A1"] = TEMPLATE_SHEET
    ws["A1"].font = Font(name="黑体", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:G1")
    ws["B2"] = "发货日期:"; ws["C2"] = ship_date
    ws["C2"].number_format = "yyyy-mm-dd"
    ws["E2"] = "运单号："; ws["F2"] = waybill_num
    for c in ["B2","C2","E2","F2"]:
        ws[c].font = Font(name="宋体", size=11)
        ws[c].alignment = Alignment(horizontal="left", vertical="center")
    ws["A4"] = "收货人(电话):"; ws["B4"] = f"{recipient_name}{recipient_phone}"
    ws["E4"] = "收货单位:"; ws["F4"] = company_name
    ws.merge_cells("B4:C4"); ws.merge_cells("E4:G4")
    ws["A5"] = "收货地址："; ws["B5"] = address
    ws.merge_cells("B5:G5")
    headers = ["品名","规格","件数","重量（kg）","批号","是否送货","托盘"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=col, value=h)
        c.font = Font(name="宋体", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    for i, item in enumerate(items):
        row = ITEMS_START_ROW + i
        for col, key in enumerate(["品名","规格","件数","重量","批号","是否送货","托盘"], 1):
            c = ws.cell(row=row, column=col, value=item.get(key,""))
            c.font = Font(name="宋体", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["H6"] = "第一联白联存底"; ws.merge_cells("H6:H10")
    ws["H12"] = "第二联红联回单"; ws.merge_cells("H12:H17")
    ws["H19"] = "第三联黄联收货人"; ws.merge_cells("H19:H23")
    for c in ["H6","H12","H19"]:
        ws[c].font = Font(name="宋体", size=9)
        ws[c].alignment = Alignment(horizontal="center", vertical="center")
    ws["A15"] = "合计:"
    r1,r2 = ITEMS_START_ROW, ITEMS_END_ROW
    ws["C15"] = f"=COUNTA(C{r1}:C{r2})"
    ws["D15"] = f"=SUM(D{r1}:D{r2})"
    ws["G15"] = f"=COUNTA(G{r1}:G{r2})"
    ws["A16"] = ("备注:(本发货单一式三联,第二联回单联为本公司收货及财务结算凭证,"
                 "请妥善保存。1吨及1吨以上均要送货。1吨以下有标注送货的要送货)")
    ws["A16"].font = Font(name="宋体", size=8)
    ws["A16"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A16:G18")
    ws["A20"] = "收货人签章:"; ws["E24"] = "收货日期:"
    ws["A25"] = ("重庆华美物流有限公司        联系人:杨卉梅"
                 "       电话:13667625772"
                 "       地址:重庆两江新区华荣货运市场A422")
    ws["A25"].font = Font(name="宋体", size=9)
    ws.merge_cells("A25:G25")
    widths = {1:22,2:10,3:14,4:10,5:12,6:10,7:10,8:16}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def gen_invoice(inp_file, ship_date, output_file):
    wb_in = openpyxl.load_workbook(inp_file)
    ws_in = wb_in.active
    rows = list(ws_in.iter_rows(values_only=True))
    header = rows[1]
    col_map = {h:i for i,h in enumerate(header) if h}
    def g(row, name):
        idx = col_map.get(name)
        return row[idx] if idx is not None else None
    groups = {}
    for row in rows[2:]:
        if not any(row): continue
        name = g(row,"收货人") or ""
        addr = g(row,"收货地址") or ""
        key = (name, addr)
        if key not in groups: groups[key] = []
        groups[key].append({
            "品名": g(row,"品名") or "", "规格": g(row,"规格") or "",
            "件数": g(row,"件数") or "", "重量": g(row,"重量") or 0,
            "批号": g(row,"批号及相应桶数") or "",
            "是否送货": g(row,"低于1吨的是否送货") or "",
            "托盘": g(row,"是否运输公司打托盘") or "",
        })
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    base = int(ship_date.replace("-","").replace("/","")) * 10000
    for si, ((name,addr), items) in enumerate(groups):
        if not items: continue
        items = [it for it in items if it["品名"]]
        pm = re.search(r"1[3-9]\d{9}", str(name))
        rname = name[:pm.start()] if pm else name
        rphone = pm.group() if pm else ""
        company = items[0].get("收货单位","") if "收货单位" in col_map else ""
        sname = f"{rname[:8]}{rphone[-11:] if rphone else ''}"
        ws = wb_out.create_sheet(sname)
        build_sheet(ws, rname, rphone, company, addr, items, base+si+1, ship_date)
    wb_out.save(output_file)

# --- Kivy UI ---
if HAVE_KIVY:
    class HuameiApp(App):
        def build(self):
            self.title = "华美物流发货单"
            root = BoxLayout(orientation="vertical", padding=20, spacing=10)
            root.add_widget(Label(text="🚚 华美物流发货单生成器",
                                  font_size="20sp", size_hint_y=None, height=50))
            root.add_widget(Label(text="（请在电脑端访问 http://电脑IP:5000 使用完整功能）",
                                  font_size="12sp", size_hint_y=None, height=30,
                                  color=(0.5,0.5,0.5,1)))
            return root

    if __name__ == "__main__":
        HuameiApp().run()
else:
    # Fallback: CLI
    if len(sys.argv) >= 3:
        gen_invoice(sys.argv[1], sys.argv[2], sys.argv[3])
    else:
        inp = input("发货文件: ").strip()
        date = input("日期(20260331): ").strip()
        gen_invoice(inp, date, f"华美物流发货单{date}.xlsx")
        print("OK")
